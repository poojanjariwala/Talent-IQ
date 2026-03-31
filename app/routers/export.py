import uuid
import io
from typing import List
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.repositories.candidate_repo import CandidateRepository
from app.repositories.job_repo import JobRepository
from app.routers.auth import get_current_recruiter

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/job/{job_id}/excel")
async def export_job_excel(
    job_id: uuid.UUID,
    recruiter=Depends(get_current_recruiter),
    db: AsyncSession = Depends(get_db),
):
    """Export all candidates for a job as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    job_repo = JobRepository(db)
    job = await job_repo.get_by_id(job_id)
    if not job or job.recruiter_id != recruiter.id:
        raise HTTPException(status_code=404, detail="Job not found")

    candidate_repo = CandidateRepository(db)
    candidates = await candidate_repo.list_by_job(job_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "Name", "Email", "Phone", "Location", "Experience (Yrs)",
        "Match Score", "Bucket", "Stage", "Matched Skills", "Missing Skills",
        "Normalized Skills", "Education", "Source File", "Parse Status", "Notes"
    ]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Bucket color map
    bucket_colors = {
        "shortlist": "d4edda",
        "hold": "fff3cd",
        "reject": "f8d7da",
        "new": "cfe2ff",
    }

    for row_idx, c in enumerate(candidates, 2):
        education_str = "; ".join(
            f"{e.get('degree', '')} @ {e.get('institution', '')}" for e in (c.education or [])
        )
        row_data = [
            c.full_name or "",
            c.email or "",
            c.phone or "",
            c.location or "",
            c.total_experience_years or "",
            f"{c.match_score:.1f}" if c.match_score is not None else "",
            c.bucket,
            c.stage,
            ", ".join(c.matched_skills or []),
            ", ".join(c.missing_skills or []),
            ", ".join(c.normalized_skills or []),
            education_str,
            c.source_filename or "",
            c.parse_status,
            c.notes or "",
        ]
        bucket_color = bucket_colors.get(c.bucket, "FFFFFF")
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=bucket_color)

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(candidates) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"candidates_{job.title.replace(' ', '_')}_{job_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
